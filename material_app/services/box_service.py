# material_app/services/box_service.py
import logging
from collections import defaultdict
from io import BytesIO

from django.db import transaction
from django.shortcuts import get_object_or_404
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..models import MaterialItems, MaterialOverview, TransactionLog

logger = logging.getLogger(__name__)


# ── 查詢 ──────────────────────────────────────────────

def get_all_boxes():
    return MaterialOverview.objects.all().order_by('-created_at')


def get_box_or_none(box_id):
    try:
        return MaterialOverview.objects.get(box_id=box_id)
    except MaterialOverview.DoesNotExist:
        return None


def box_exists(box_id):
    return MaterialOverview.objects.filter(box_id=box_id).exists()


# ── 建立 ──────────────────────────────────────────────

def create_box(box_id, box_type, description, owner, status_val, is_locked=False):
    """
    舊版：owner 存字串 id → 新版：owner 傳 User 物件（FK）
    舊版：category → 新版：box_type（ENUM）
    """
    return MaterialOverview.objects.create(
        box_id=box_id,
        box_type=box_type or None,
        description=description or None,
        owner=owner,                    # User 物件
        status=status_val or None,
        is_locked=is_locked,
    )


# ── 更新 ──────────────────────────────────────────────

def update_box(box, box_type, description, owner, status_val, is_locked):
    box.box_type    = box_type or None
    box.description = description or None
    box.owner       = owner or None
    box.status      = status_val or None
    box.is_locked   = is_locked
    box.save()
    return box


def toggle_box_lock(box, action):
    if action == 'lock':
        box.is_locked = True
        box.locked_at = timezone.now()
    elif action == 'unlock':
        box.is_locked = False
        box.locked_at = timezone.now()
    else:
        raise ValueError("action 必須是 lock 或 unlock")
    box.save()
    return box


# ── 刪除 ──────────────────────────────────────────────

def delete_box(box):
    # 舊版：filter(BoxID=box) → 新版：filter(box=box)
    items_count = MaterialItems.objects.filter(box=box).count()
    MaterialItems.objects.filter(box=box).delete()
    box.delete()
    return items_count


# ── 入庫 ──────────────────────────────────────────────

@transaction.atomic
def checkin_items(box_id, selected_items, qty_map, operator):
    """
    將物品入庫到指定容器。

    舊版：operator_id 傳字串
    新版：operator 傳 User 物件（FK）
    舊版：無 select_for_update
    新版：加上 select_for_update 防並發
    """
    target_box = get_object_or_404(MaterialOverview, box_id=box_id)

    if target_box.is_locked:
        raise ValueError(f'容器 {box_id} 已鎖定，無法入庫')

    success_count = 0
    for sn in selected_items:
        qty = qty_map.get(sn, 0)
        if qty <= 0:
            continue

        source_item     = MaterialItems.objects.select_for_update().get(sn=sn)
        original_box_id = source_item.box.box_id
        stock_before    = source_item.quantity

        if source_item.box == target_box:
            # 同箱直接加數量
            source_item.quantity += qty
            source_item.save()
            action       = 'IN'
            remark       = '從容器列表直接入庫'
            current_item = source_item
        else:
            # 跨箱：目的箱加、來源箱扣
            target_item, _ = MaterialItems.objects.select_for_update().get_or_create(
                sn=sn, box=target_box,
                defaults={
                    'item_name': source_item.item_name,
                    'spec':      source_item.spec,
                    'location':  source_item.location,
                    'category':  source_item.category,
                    'quantity':  0,
                }
            )
            target_item.quantity += qty
            target_item.save()

            source_item.quantity -= qty
            if source_item.quantity <= 0:
                source_item.delete()
            else:
                source_item.save()

            action       = 'MOVE'
            remark       = f'入庫調撥（來源：{original_box_id}）'
            current_item = target_item

        TransactionLog.objects.create(
            action_type=action,
            item=current_item,
            from_box_id=original_box_id if action == 'MOVE' else None,
            to_box_id=box_id,
            trans_qty=qty,
            stock_before=stock_before,
            stock_after=current_item.quantity,
            operator=operator,          # 舊版字串 → 新版 User FK
            remark=remark,
        )
        success_count += 1

    return success_count


# ── BOM ───────────────────────────────────────────────

def get_bom_data(box):
    # 舊版：filter(BoxID=box) → 新版：filter(box=box)
    all_items       = MaterialItems.objects.filter(box=box).order_by('sn')
    bom_items       = [item for item in all_items if item.is_bom_item]
    total           = len(bom_items)
    fulfilled_count = sum(1 for item in bom_items if item.bom_status == 'fulfilled')
    is_all_ready    = total > 0 and total == fulfilled_count
    existing_items  = list(MaterialItems.objects.values('sn', 'item_name', 'spec').distinct().order_by('sn'))
    return {
        'all_items':       all_items,
        'bom_items':       bom_items,
        'total':           total,
        'fulfilled_count': fulfilled_count,
        'is_all_ready':    is_all_ready,
        'existing_items':  existing_items,
    }


def get_bom_summary(box):
    """BOM 進度摘要，供 project_list 顯示進度條用。"""
    all_items  = MaterialItems.objects.filter(box=box)
    bom_items  = [item for item in all_items if item.is_bom_item]
    total      = len(bom_items)
    fulfilled  = sum(1 for item in bom_items if item.bom_status == 'fulfilled')
    pct        = int(fulfilled / total * 100) if total > 0 else 0
    return {'total': total, 'fulfilled': fulfilled, 'pct': pct}


@transaction.atomic
def pickup_bom(box, operator):
    """
    領料：把容器內所有 BOM 物料按 required_qty 扣減庫存。

    舊版：operator_id 字串、RequiredQty、BoxID 大寫
    新版：operator User FK、required_qty、box_id snake_case
    """
    all_items    = MaterialItems.objects.select_for_update().filter(box=box).order_by('sn')
    bom_items    = [item for item in all_items if item.is_bom_item]
    is_all_ready = all(item.bom_status == 'fulfilled' for item in bom_items) and len(bom_items) > 0

    if not is_all_ready:
        raise ValueError('物料尚未全數到齊，無法執行領料')

    count = 0
    for item in bom_items:
        stock_before    = item.quantity
        item.quantity  -= item.required_qty
        item.save()

        TransactionLog.objects.create(
            action_type='OUT',
            item=item,
            from_box_id=box.box_id,
            trans_qty=item.required_qty,
            stock_before=stock_before,
            stock_after=item.quantity,
            operator=operator,
            remark=f'BOM 領料 - {box.box_id}',
        )
        count += 1

    return count


# ── Excel ─────────────────────────────────────────────

def export_boxes_to_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "容器列表"

    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['容器編號', '類型', '描述', '負責人', '狀態', '是否鎖定', '建立日期']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border

    boxes = MaterialOverview.objects.all().order_by('box_id')
    for box in boxes:
        ws.append([
            box.box_id,
            box.box_type or '',
            box.description or '',
            # owner 現在是 FK，取 username；若無則空字串
            box.owner.username if box.owner else '',
            box.status or '',
            '是' if box.is_locked else '否',
            box.created_at.strftime('%Y-%m-%d %H:%M:%S') if box.created_at else '',
        ])
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border    = border

    for i, width in enumerate([15, 12, 30, 12, 12, 12, 20], 1):
        ws.column_dimensions[chr(64 + i)].width = width

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file, boxes.count()


def import_boxes_from_excel(excel_file, current_user):
    """
    舊版：owner 存字串 id → 新版：owner 傳 User 物件
    Excel 裡的 owner 欄位改成 username，查 User 表取得物件。
    """
    from django.contrib.auth.models import User

    wb = load_workbook(excel_file)
    ws = wb.active

    success_count = skip_count = error_count = 0
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if not any(row):
                continue

            box_id = str(row[0]).strip() if row[0] else None
            if not box_id:
                errors.append(f'第 {row_num} 行：容器編號不能為空')
                error_count += 1
                continue

            if MaterialOverview.objects.filter(box_id=box_id).exists():
                skip_count += 1
                continue

            box_type    = str(row[1]).strip() if row[1] else None
            description = str(row[2]).strip() if row[2] else None
            status_val  = str(row[4]).strip() if len(row) > 4 and row[4] else '使用中'
            locked_val  = str(row[5]).strip() if len(row) > 5 and row[5] else '否'
            is_locked   = locked_val in ['是', 'True', 'true', '1', 'YES', 'yes']

            # owner：Excel 填 username，查不到則用匯入者
            owner_username = str(row[3]).strip() if row[3] else None
            if owner_username:
                owner = User.objects.filter(username=owner_username).first() or current_user
            else:
                owner = current_user

            MaterialOverview.objects.create(
                box_id=box_id,
                box_type=box_type,
                description=description,
                owner=owner,
                status=status_val,
                is_locked=is_locked,
            )
            success_count += 1

        except Exception as e:
            errors.append(f'第 {row_num} 行：{str(e)}')
            error_count += 1

    return success_count, skip_count, error_count, errors


def build_download_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "容器匯入範本"

    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['容器編號*', '類型', '描述', '負責人(username)', '狀態', '是否鎖定']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border

    for row_data in [
        ['BOX-001', 'project', '測試專案容器', 'alice', '使用中', '否'],
        ['BOX-002', 'personal', '個人物品',    'bob',   '空閒',   '否'],
        ['BOX-003', 'warehouse', '倉庫備品',   'carol', '使用中', '是'],
    ]:
        ws.append(row_data)
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border    = border

    ws.append([])
    for note in [
        ['說明：'], ['1. 帶 * 的欄位為必填'], ['2. 容器編號不可重複'],
        ['3. 負責人填 username，查不到則自動設定為匯入者'],
        ['4. 狀態可填：使用中、空閒、結案'],
        ['5. 是否鎖定可填：是、否'],
        ['6. 類型可填：project / personal / warehouse / spare / test / other'],
    ]:
        ws.append(note)

    for i, width in enumerate([15, 12, 30, 20, 12, 12], 1):
        ws.column_dimensions[chr(64 + i)].width = width

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file


# ── 統計 ──────────────────────────────────────────────

def get_dashboard_stats():
    try:
        return {
            'total_boxes':         MaterialOverview.objects.count(),
            'total_items':         MaterialItems.objects.count(),
            # 舊版：filter(Locked=True) → 新版：filter(is_locked=True)
            'locked_boxes':        MaterialOverview.objects.filter(is_locked=True).count(),
            'recent_transactions': TransactionLog.objects.count(),
        }
    except Exception as e:
        logger.error(f"獲取統計數據失敗: {str(e)}")
        return {'total_boxes': 0, 'total_items': 0, 'locked_boxes': 0, 'recent_transactions': 0}